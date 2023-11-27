
from bs4 import BeautifulSoup
import urllib.request               
from time import sleep
from datetime import datetime
import pandas as pd
import requests
import re
from datetime import date
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sb



def getweather():
    weather = []
    year = date.today().year
    url = "http://www.hko.gov.hk/cis/dailyExtract/dailyExtract_" + str(year) + "08.xml"

    page = requests.get(url)

    soup = BeautifulSoup(page.content, 'lxml')

    body = soup.find("body").text
    body = body.split(",")
    # weather.append(temp)
    day = body[-27][4:-1]
    if (day[0] == "0"):
        day_2 = day[-1]
    else:
        day_2 = day
    weather.append(day_2)
    month = body[0][-1]
    weather.append(month)
    weather.append(year)

    temp = body[-24][1:-1]
    weather.append(temp)

    high = body[-25][1:-1]

    low = body[-23][1:-1]
    weather.append(high)
    weather.append(low)

    Humidity = body[-21][1:-1]
    weather.append(Humidity)
    dew = body[-22][1:-1]
    weather.append(dew)
    pressure = body[-26][1:-1]
    weather.append(pressure)
    if (float(temp) >= 30):
        heat = "YES"
    else:
        heat = "NO"
    weather.append(heat)
    if (float(Humidity) >= 80):
        wet = "YES"
    else:
        wet = "NO"
    weather.append(wet)

    print("Ngày : ", day_2)
    print("Tháng : ", body[0][-1])
    print("Năm : ", year)
    print("Áp suất : ", body[-26][1:-1])
    print("Nhiệt độ Max : ", body[-25][1:-1])
    print("Nhiệt độ trung bình : ", body[-24][1:-1])
    print("Nhiệt độ Min : ", body[-23][1:-1])
    print("Điểm sương : ", body[-22][1:-1])
    print("Độ ẩm : ", body[-21][1:-1])
    print("Nhiệt : ", heat)
    print("Ẩm ướt : ", wet)
    # print("Temperature : ",temp)

    return weather


getweather()


import csv
import openpyxl
import sys



def main():
    print("Thu thập dữ liệu thời tiết trong khoảng thời gian 30 phút")
    idx = 0
    weatherdata = {'Day':[],'Month':[],'Year':[],'mean_temp':[],'max_temp':[],'min_temp':[],'meanhum':[],'meandew':[],'pressure':[],'heat':[],'wet':[]}
    filename = 'data.xlsx'
    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb['Sheet1']
    #new_row = ['Day','Month','Year','mean_temp','max_temp','min_temp','meanhum','meandew','pressure','heat','wet']
    #sheet.append(new_row)
    while idx < 1:
        print('collecting weather data '+str(idx))
        tmp = getweather()
        weatherdata['Day'].append(tmp[0])
        weatherdata['Month'].append(tmp[1])
        weatherdata['Year'].append(tmp[2])
        weatherdata['mean_temp'].append(tmp[3])
        weatherdata['max_temp'].append(tmp[4])
        weatherdata['min_temp'].append(tmp[5])
        weatherdata['meanhum'].append(tmp[6])
        weatherdata['meandew'].append(tmp[7])
        weatherdata['pressure'].append(tmp[8])
        weatherdata['heat'].append(tmp[9])
        weatherdata['wet'].append(tmp[10])
        
        try :
            new_row = [int(tmp[0]),int(tmp[1]),int(tmp[2]),float(tmp[3]),float(tmp[4]),float(tmp[5]),float(tmp[6]),float(tmp[7]),float(tmp[8]),tmp[9],tmp[10]]
            sheet.append(new_row)
        except ValueError as e:
            print (e)
        idx+=1

    wb.save(filename)


    print(weatherdata)


main()

data_xls = pd.read_excel('data.xlsx', 'Sheet1', index_col=None)
data_xls.to_csv('data.csv', encoding='utf-8', index=False)


#importing classes to handle data,to split data into training and testing sets,to visualising of tree and to cheak accuracy of our model
from preprocessing import splitter
from preprocessing import Encoder
from preprocessing import scaler
from metrics import matrix
from treeVis import vis
sp=splitter()
mt=matrix()
vs=vis()
sc=scaler()
en=Encoder()

#importing dataset
dataset=pd.read_csv("data.csv")
dataset.tail()


from sklearn.preprocessing import LabelEncoder
lab=LabelEncoder()
dataset.iloc[:,0]=lab.fit_transform(dataset.iloc[:,0])
dataset.iloc[:,1]=lab.fit_transform(dataset.iloc[:,1])
dataset.iloc[:,2]=lab.fit_transform(dataset.iloc[:,2])
dataset.iloc[:,3]=lab.fit_transform(dataset.iloc[:,3])
dataset.iloc[:,4]=lab.fit_transform(dataset.iloc[:,4])
dataset.iloc[:,5]=lab.fit_transform(dataset.iloc[:,5])
dataset.iloc[:,6]=lab.fit_transform(dataset.iloc[:,6])
dataset.iloc[:,7]=lab.fit_transform(dataset.iloc[:,7])
dataset.iloc[:,8]=lab.fit_transform(dataset.iloc[:,8])
dataset.iloc[:,9]=lab.fit_transform(dataset.iloc[:,9])
dataset.iloc[:,10]=lab.fit_transform(dataset.iloc[:,10])

dataset.tail()

#deviding data into dependant and independant sets
x = dataset[['Day','Month','Year']]
y = dataset['heat']
z = dataset['wet']

#visualising each attribute of dataset using histogram
dataset.hist(figsize = (30, 30))
plt.savefig("dataset.png")
plt.show()


#deviding data into training and testing sets
from sklearn.model_selection import train_test_split
x_train,x_test,y_train,y_test=train_test_split(x,y,test_size=0.20,random_state=42)
x_train,x_test,z_train,z_test=train_test_split(x,z,test_size=0.20,random_state=42)


#importing decision tree model and fitting training data to it
from sklearn.tree import DecisionTreeClassifier
classifier=DecisionTreeClassifier(criterion='entropy',random_state=0)
classifier.fit(x,y)

classifier_z=DecisionTreeClassifier(criterion='entropy',random_state=0)
classifier_z.fit(x,z)


#predicting values for x_test and compairing result with y_test
print("Heat:-")
y_pred=classifier.predict(x_test)
print("predicted values:-")
print(y_pred.astype(int))
y_test_arr=np.array(y_test)
print("original values:-")
print(y_test_arr)

#predicting values for x_test and compairing result with y_test
print("Wet:-")
z_pred=classifier_z.predict(x_test)
print("predicted values:-")
print(z_pred.astype(int))
z_test_arr=np.array(z_test)
print("original values:-")
print(z_test_arr)


#cheaking accuracy of our model
accuracy=mt.accuracy(y_pred,y_test)
print("{}%".format(accuracy*100))

accuracy=mt.accuracy(z_pred,z_test)
print("{}%".format(accuracy*100))

from datetime import date
now = date.today()
filepath = "prediction/forecast-"+str(now)+".xlsx"
def prediction():
    wb = openpyxl.Workbook()
    wb.save(filepath)
    print("Dự Đoán Thông Tin Thời Tiết ")
    idx = 0
    weatherdata = {'Day':[],'Month':[],'Year':[],'heat':[],'wet':[]}
    filename = filepath
    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb['Sheet']
    new_row = ['Day','Month','Year','Heat','Wet']
    sheet.append(new_row)
    while idx < 31:
        print('Dự đoán dữ liệu thời tiết ' + str(idx))
        year = date.today().year
        yr = year - 1999
        if (date.today().day == 31):
            year += 1
            yr += 1
        y_pred = classifier.predict([[idx, 0, yr]])
        print(y_pred.astype(int))
        z_pred = classifier_z.predict([[idx, 0, yr]])
        print(z_pred.astype(int))
        tmp_d = idx + 1
        tmp_m = 8
        tmp_y = year
        tmp_h = y_pred.astype(int)
        tmp_w = z_pred.astype(int)
        if (int(tmp_h) == 0):
            tmp_h_char = "NO"
        else:
            tmp_h_char = "YES"
        if (int(tmp_w) == 0):
            tmp_w_char = "NO"
        else:
            tmp_w_char = "YES"
        new_row = [int(tmp_d), int(tmp_m), int(tmp_y), tmp_h_char, tmp_w_char]
        sheet.append(new_row)
        idx += 1
    wb.save(filename)
    
    print(weatherdata)


prediction()


data_xls = pd.read_excel(filepath, 'Sheet', index_col=None)
data_xls.to_csv("prediction/forecast-"+str(now)+".csv", encoding='utf-8', index=False)

#importing dataset
data_pred=pd.read_csv("prediction/forecast-"+str(now)+".csv")
data_pred


import seaborn as sns
sns_plot = sns.countplot(x='Heat',data=data_pred)
fig = sns_plot.get_figure()
fig.savefig("prediction/heat_summary_prediction-"+str(now)+".png")


sns_plot = sns.countplot(x='Wet',data=data_pred)
fig = sns_plot.get_figure()
fig.savefig("prediction/wet_summary_prediction-"+str(now)+".png")


